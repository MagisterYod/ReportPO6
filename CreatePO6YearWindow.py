import tkinter as tk
import time
import openpyxl
import os
import threading
from variables import p_date, p_time, path, sql_shovs_year, sql_veh_year, sql_bul_year, sql_pogr_year
from tkinter import ttk, Button, Checkbutton, messagebox, W


class CreateWindowPO6Year(ttk.Frame):

    def __init__(self, master):
        super().__init__(master)
        self.jan = tk.IntVar()
        self.feb = tk.IntVar()
        self.mar = tk.IntVar()
        self.apr = tk.IntVar()
        self.may = tk.IntVar()
        self.jun = tk.IntVar()
        self.jul = tk.IntVar()
        self.aug = tk.IntVar()
        self.sep = tk.IntVar()
        self.oct = tk.IntVar()
        self.nov = tk.IntVar()
        self.dec = tk.IntVar()
        self.quarter_1 = tk.BooleanVar()
        self.quarter_2 = tk.BooleanVar()
        self.quarter_3 = tk.BooleanVar()
        self.quarter_4 = tk.BooleanVar()
        self.year = tk.BooleanVar()
        self.jan_checkbox = Checkbutton(self, text="Январь", onvalue=1, offvalue=False, variable=self.jan)
        self.feb_checkbox = Checkbutton(self, text="Февраль", onvalue=2, offvalue=False, variable=self.feb)
        self.mar_checkbox = Checkbutton(self, text="Март", onvalue=3, offvalue=False, variable=self.mar)
        self.apr_checkbox = Checkbutton(self, text="Апрель", onvalue=4, offvalue=False, variable=self.apr)
        self.may_checkbox = Checkbutton(self, text="Май", onvalue=5, offvalue=False, variable=self.may)
        self.jun_checkbox = Checkbutton(self, text="Июнь", onvalue=6, offvalue=False, variable=self.jun)
        self.jul_checkbox = Checkbutton(self, text="Июль", onvalue=7, offvalue=False, variable=self.jul)
        self.aug_checkbox = Checkbutton(self, text="Август", onvalue=8, offvalue=False, variable=self.aug)
        self.sep_checkbox = Checkbutton(self, text="Сентябрь", onvalue=9, offvalue=False, variable=self.sep)
        self.oct_checkbox = Checkbutton(self, text="Октябрь", onvalue=10, offvalue=False, variable=self.oct)
        self.nov_checkbox = Checkbutton(self, text="Ноябрь", onvalue=11, offvalue=False, variable=self.nov)
        self.dec_checkbox = Checkbutton(self, text="Декабрь", onvalue=12, offvalue=False, variable=self.dec)
        self.first_quarter = Checkbutton(self, text='Первый квартал', variable=self.quarter_1)
        self.second_quarter = Checkbutton(self, text='Второй квартал', variable=self.quarter_2)
        self.third_quarter = Checkbutton(self, text='Третий квартал', variable=self.quarter_3)
        self.fourth_quarter = Checkbutton(self, text='Четвертый квартал', variable=self.quarter_4)
        self.all_year = Checkbutton(self, text='За весь год', variable=self.year, font="arial 10 bold")
        self.entry = ttk.Spinbox(self, from_=2000, to=2100, justify='center', wrap=True, width=10)
        self.entry.insert(0, p_date[2])
        self.start_button = Button(self, text='Выгрузить отчет', command=self.start, fg='green', width=42)
        self.text_progress_wb = ttk.Label(self, text='Месяц:', state='disable', width=42)
        self.progress_wb = ttk.Progressbar(self, orient='horizontal', mode='determinate', length=304, maximum=300)
        self.text_progress_ws = ttk.Label(self, text='Тип техники:', state='disable', width=42)
        self.text_progress_row = ttk.Label(self, text='Техника:', state='disable', width=42)
        self.po6_year_menu()

    def po6_year_menu(self):
        self.jan_checkbox.grid(row=0, column=0, sticky=W)
        self.feb_checkbox.grid(row=1, column=0, sticky=W)
        self.mar_checkbox.grid(row=2, column=0, sticky=W)
        self.apr_checkbox.grid(row=3, column=0, sticky=W)
        self.may_checkbox.grid(row=4, column=0, sticky=W)
        self.jun_checkbox.grid(row=5, column=0, sticky=W)
        self.jul_checkbox.grid(row=6, column=0, sticky=W)
        self.aug_checkbox.grid(row=7, column=0, sticky=W)
        self.sep_checkbox.grid(row=8, column=0, sticky=W)
        self.oct_checkbox.grid(row=9, column=0, sticky=W)
        self.nov_checkbox.grid(row=10, column=0, sticky=W)
        self.dec_checkbox.grid(row=11, column=0, sticky=W)
        self.first_quarter.grid(row=0, column=1, rowspan=3, sticky=W)
        self.second_quarter.grid(row=3, column=1, rowspan=3, sticky=W)
        self.third_quarter.grid(row=6, column=1, rowspan=3, sticky=W)
        self.fourth_quarter.grid(row=9, column=1, rowspan=3, sticky=W)
        self.all_year.grid(row=12, column=0, sticky=W)
        self.entry.grid(row=12, column=1, sticky=W)
        self.start_button.grid(row=13, column=0, columnspan=2, pady=5)

    def quarters(self):
        if self.year.get():
            self.quarter_1.set(True)
            self.quarter_2.set(True)
            self.quarter_3.set(True)
            self.quarter_4.set(True)
        if self.quarter_1.get():
            self.jan.set(1)
            self.feb.set(2)
            self.mar.set(3)
        if self.quarter_2.get():
            self.apr.set(4)
            self.may.set(5)
            self.jun.set(6)
        if self.quarter_3.get():
            self.jul.set(7)
            self.aug.set(8)
            self.sep.set(9)
        if self.quarter_4.get():
            self.oct.set(10)
            self.nov.set(11)
            self.dec.set(12)

    def refresh(self):
        self.master.update()
        self.master.after(1000, self.refresh)

    def start(self):
        self.refresh()
        threading.Thread(target=self.take_doc).start()

    def timer_start(self):
        self.jan_checkbox.grid_forget()
        self.feb_checkbox.grid_forget()
        self.mar_checkbox.grid_forget()
        self.apr_checkbox.grid_forget()
        self.may_checkbox.grid_forget()
        self.jun_checkbox.grid_forget()
        self.jul_checkbox.grid_forget()
        self.aug_checkbox.grid_forget()
        self.sep_checkbox.grid_forget()
        self.oct_checkbox.grid_forget()
        self.nov_checkbox.grid_forget()
        self.dec_checkbox.grid_forget()
        self.first_quarter.grid_forget()
        self.second_quarter.grid_forget()
        self.third_quarter.grid_forget()
        self.fourth_quarter.grid_forget()
        self.all_year.grid_forget()
        self.entry.grid_forget()
        self.start_button.grid_forget()
        self.progress_wb.grid(row=14, column=0, columnspan=3)
        self.text_progress_wb.grid(row=15, column=0, columnspan=3, sticky=W)
        self.text_progress_ws.grid(row=16, column=0, columnspan=3, sticky=W)
        self.text_progress_row.grid(row=17, column=0, columnspan=3, sticky=W)
        self.update_idletasks()

    def timer_stop(self):
        self.master.destroy()

    def changes(self, month, technics, tech, percent, sec):
        self.text_progress_wb['text'] = f'Месяц: {month}'
        self.text_progress_ws['text'] = f'Тип техники: {technics}'
        if tech == '':
            self.text_progress_row['text'] = f'Идёт выгрузка данных из базы данных...'
        else:
            self.text_progress_row['text'] = f'Техника: {tech}'
        self.progress_wb['value'] += percent
        self.update_idletasks()
        time.sleep(sec)

    def take_doc(self):

        def insert_technics(type_tech, row_tech, df):
            if type_tech == 'shov' and len(df) == 34:
                ws[f'F{row_tech}'] = df[1]  # 6 ДВС
                ws[f'K{row_tech}'] = df[2]  # 11 объем работ план
                ws[f'L{row_tech}'] = df[3]  # 12 объем работ факт
                ws[f'M{row_tech}'] = df[4]  # 13 план ТО
                ws[f'N{row_tech}'] = df[5]  # 14 факт ТО
                ws[f'O{row_tech}'] = df[6]  # 15 план ТР
                ws[f'P{row_tech}'] = df[7]  # 16 факт ТР
                ws[f'Q{row_tech}'] = df[8]  # 17 план КР
                ws[f'R{row_tech}'] = df[9]  # 18 фатк КР
                ws[f'S{row_tech}'] = df[10]  # 19 план регламент
                ws[f'T{row_tech}'] = df[11]  # 20 факт регламент
                ws[f'U{row_tech}'] = df[12]  # 21 план обед
                ws[f'V{row_tech}'] = df[13]  # 22 факт обед
                ws[f'X{row_tech}'] = df[14]  # 24 факт прием/передача
                ws[f'Y{row_tech}'] = df[15]  # 25 план забой
                ws[f'Z{row_tech}'] = df[16]  # 26 факт забой
                ws[f'AA{row_tech}'] = df[17]  # 27 план БВР
                ws[f'AB{row_tech}'] = df[18]  # 28 факт БВР
                ws[f'AC{row_tech}'] = df[19]  # 29 ДВС
                ws[f'AD{row_tech}'] = df[20]  # 30 трансмиссия
                ws[f'AE{row_tech}'] = df[21]  # 31 ходовая
                ws[f'AF{row_tech}'] = df[22]  # 32 навесное
                ws[f'AG{row_tech}'] = df[23]  # 33 электро
                ws[f'AH{row_tech}'] = df[24]  # 34 гибравлика
                ws[f'AI{row_tech}'] = df[25]  # 35 прочие
                ws[f'AJ{row_tech}'] = df[26]  # 36 автошины
                ws[f'AK{row_tech}'] = df[27]  # 37 фронт работ
                ws[f'AL{row_tech}'] = df[28]  # 38 зап.части
                ws[f'AM{row_tech}'] = df[29]  # 39 ГСМ
                ws[f'AN{row_tech}'] = df[30]  # 40 персонал
                ws[f'AO{row_tech}'] = df[31]  # 41 метеоусловия
                ws[f'AP{row_tech}'] = df[32]  # 42 прочие
                ws[f'AT{row_tech}'] = int(df[33])  # 46 фонд времени
            if type_tech == 'veh' and len(df) == 34:
                ws[f'F{row_tech}'] = df[1]  # 6 ДВС
                ws[f'H{row_tech}'] = df[2]  # 8 пробег общий
                ws[f'I{row_tech}'] = df[3]  # 9 пробег с грузом
                ws[f'K{row_tech}'] = df[4]  # 11 объем работ план
                ws[f'L{row_tech}'] = df[5]  # 12 объем работ факт
                ws[f'M{row_tech}'] = df[6]  # 13 план ТО
                ws[f'N{row_tech}'] = df[7]  # 14 факт ТО
                ws[f'O{row_tech}'] = df[8]  # 15 план ТР
                ws[f'P{row_tech}'] = df[9]  # 16 факт ТР
                ws[f'Q{row_tech}'] = df[10]  # 17 план КР
                ws[f'R{row_tech}'] = df[11]  # 18 фатк КР
                ws[f'S{row_tech}'] = df[12]  # 19 план регламент
                ws[f'T{row_tech}'] = df[13]  # 20 факт регламент
                ws[f'U{row_tech}'] = df[14]  # 21 план обед
                ws[f'V{row_tech}'] = df[15]  # 22 факт обед
                ws[f'X{row_tech}'] = df[16]  # 24 факт прием/передача
                ws[f'AA{row_tech}'] = df[17]  # 27 план БВР
                ws[f'AB{row_tech}'] = df[18]  # 28 факт БВР
                ws[f'AC{row_tech}'] = df[19]  # 29 ДВС
                ws[f'AD{row_tech}'] = df[20]  # 30 трансмиссия
                ws[f'AE{row_tech}'] = df[21]  # 31 ходовая
                ws[f'AF{row_tech}'] = df[22]  # 32 навесное
                ws[f'AG{row_tech}'] = df[23]  # 33 электро
                ws[f'AH{row_tech}'] = df[24]  # 34 гибравлика
                ws[f'AI{row_tech}'] = df[25]  # 35 прочие
                ws[f'AJ{row_tech}'] = df[26]  # 36 автошины
                ws[f'AK{row_tech}'] = df[27]  # 37 фронт работ
                ws[f'AL{row_tech}'] = df[28]  # 38 зап.части
                ws[f'AM{row_tech}'] = df[29]  # 39 ГСМ
                ws[f'AN{row_tech}'] = df[30]  # 40 персонал
                ws[f'AO{row_tech}'] = df[31]  # 41 метеоусловия
                ws[f'AP{row_tech}'] = df[32]  # 42 прочие
                ws[f'AT{row_tech}'] = int(df[33])  # 46 фонд времени
            if type_tech in ['bul', 'pogr'] and len(df) == 30:
                ws[f'F{row_tech}'] = df[1]  # 6 ДВС
                ws[f'M{row_tech}'] = df[2]  # 13 план ТО
                ws[f'N{row_tech}'] = df[3]  # 14 факт ТО
                ws[f'O{row_tech}'] = df[4]  # 15 план ТР
                ws[f'P{row_tech}'] = df[5]  # 16 факт ТР
                ws[f'Q{row_tech}'] = df[6]  # 17 план КР
                ws[f'R{row_tech}'] = df[7]  # 18 фатк КР
                ws[f'S{row_tech}'] = df[8]  # 19 план регламент
                ws[f'T{row_tech}'] = df[9]  # 20 факт регламент
                ws[f'U{row_tech}'] = df[10]  # 21 план обед
                ws[f'V{row_tech}'] = df[11]  # 22 факт обед
                ws[f'X{row_tech}'] = df[12]  # 24 факт прием/передача
                ws[f'AA{row_tech}'] = df[13]  # 27 план БВР
                ws[f'AB{row_tech}'] = df[14]  # 28 факт БВР
                ws[f'AC{row_tech}'] = df[15]  # 29 ДВС
                ws[f'AD{row_tech}'] = df[16]  # 30 трансмиссия
                ws[f'AE{row_tech}'] = df[17]  # 31 ходовая
                ws[f'AF{row_tech}'] = df[18]  # 32 навесное
                ws[f'AG{row_tech}'] = df[19]  # 33 электро
                ws[f'AH{row_tech}'] = df[20]  # 34 гибравлика
                ws[f'AI{row_tech}'] = df[21]  # 35 прочие
                ws[f'AJ{row_tech}'] = df[22]  # 36 автошины
                ws[f'AK{row_tech}'] = df[23]  # 37 фронт работ
                ws[f'AL{row_tech}'] = df[24]  # 38 зап.части
                ws[f'AM{row_tech}'] = df[25]  # 39 ГСМ
                ws[f'AN{row_tech}'] = df[26]  # 40 персонал
                ws[f'AO{row_tech}'] = df[27]  # 41 метеоусловия
                ws[f'AP{row_tech}'] = df[28]  # 42 прочие
                ws[f'AT{row_tech}'] = int(df[29])  # 46 фонд времени

        try:
            result = []
            self.quarters()
            months = {
                f'{self.jan_checkbox["text"]}': f'01.0{self.jan.get()}.{self.entry.get()}',
                f'{self.feb_checkbox["text"]}': f'01.0{self.feb.get()}.{self.entry.get()}',
                f'{self.mar_checkbox["text"]}': f'01.0{self.mar.get()}.{self.entry.get()}',
                f'{self.apr_checkbox["text"]}': f'01.0{self.apr.get()}.{self.entry.get()}',
                f'{self.may_checkbox["text"]}': f'01.0{self.may.get()}.{self.entry.get()}',
                f'{self.jun_checkbox["text"]}': f'01.0{self.jun.get()}.{self.entry.get()}',
                f'{self.jul_checkbox["text"]}': f'01.0{self.jul.get()}.{self.entry.get()}',
                f'{self.aug_checkbox["text"]}': f'01.0{self.aug.get()}.{self.entry.get()}',
                f'{self.sep_checkbox["text"]}': f'01.0{self.sep.get()}.{self.entry.get()}',
                f'{self.oct_checkbox["text"]}': f'01.{self.oct.get()}.{self.entry.get()}',
                f'{self.nov_checkbox["text"]}': f'01.{self.nov.get()}.{self.entry.get()}',
                f'{self.dec_checkbox["text"]}': f'01.{self.dec.get()}.{self.entry.get()}'
            }

            for item in months.items():
                if item[1].split('.')[1] not in ['00', '0']:
                    result.append(item)
            print(result)
            try:
                wb = openpyxl.load_workbook('C:\\asd\\Report\\PO6_mounth_new.xlsx')
            except Exception as e:
                print(e)
            percent = round(100 / (result.__len__() * 14), 3) * 3
            self.timer_start()
            for month in result:
                self.update_idletasks()
                p_month = ''
                if month[1].split('.')[1].startswith('0'):
                    p_month = month[1].split('.')[1].lstrip('0')
                else:
                    p_month = month[1].split('.')[1]
                ws = wb[p_month]
                ws['A9'] = f'за {month[0]} {month[1].split(".")[2]}'
                self.changes(month[0], 'Экскаваторы', '', 0, 1)
                for row in sql_shovs_year(month[1]).getvalue().fetchall():
                    if row[0] == '' or row[0] is None:
                        continue
                    if int(row[0]) == 17:
                        self.changes(month[0], 'Экскаваторы', 'Шагающий 6/45 №17', percent, 3)
                        insert_technics('shov', 23, row)
                self.changes(month[0], 'Самосвалы', '', 0, 1)
                for row in sql_veh_year(month[1]).getvalue().fetchall():
                    if int(row[0]) == 32:
                        self.changes(month[0], 'Самосвалы', 'Самосвал №32', percent, 3)
                        insert_technics('veh', 26, row)
                    if int(row[0]) == 35:
                        self.changes(month[0], 'Самосвалы', 'Самосвал №35', percent, 3)
                        insert_technics('veh', 27, row)
                self.changes(month[0], 'Бульдозеры', '', 0, 1)
                for row in sql_bul_year(month[1]).getvalue().fetchall():
                    if row[0] == '' or row[0] is None:
                        continue
                    if int(row[0]) == 25:
                        self.changes(month[0], 'Бульдозеры', 'Трактор МТЗ-82 №25', percent, 3)
                        insert_technics('bul', 43, row)
                    if int(row[0]) == 38:
                        self.changes(month[0], 'Бульдозеры', 'Будьдозер ДТ-75А №38', percent, 3)
                        insert_technics('bul', 45, row)
                    if int(row[0]) == 28:
                        self.changes(month[0], 'Бульдозеры', 'Будьдозер CAT-D6R №28', percent, 3)
                        insert_technics('bul', 47, row)
                    if int(row[0]) == 33:
                        self.changes(month[0], 'Бульдозеры', 'Будьдозер CAT-D9R №33', percent, 3)
                        insert_technics('bul', 48, row)
                    if int(row[0]) == 4:
                        self.changes(month[0], 'Бульдозеры', 'Автогрейдер CAT-140М №4', percent, 3)
                        insert_technics('bul', 53, row)
                    if int(row[0]) == 5:
                        self.changes(month[0], 'Бульдозеры', 'Автогрейдер Komatsu GD825А-2 №5', percent, 3)
                        insert_technics('bul', 54, row)
                self.changes(month[0], 'Погрузчики', '', 0, 1)
                for row in sql_pogr_year(month[1]).getvalue().fetchall():
                    if row[0] == '' or row[0] is None:
                        continue
                    if int(row[0]) == 15:
                        self.changes(month[0], 'Погрузчики', 'Погрузчик CAT-962 №15', percent, 3)
                        insert_technics('pogr', 56, row)
                    if int(row[0]) == 16:
                        self.changes(month[0], 'Погрузчики', 'Погрузчик CAT-988H №16', percent, 3)
                        insert_technics('pogr', 57, row)
                    if int(row[0]) == 26:
                        self.changes(month[0], 'Погрузчики', 'Погрузчик Shantui SL60W №26', percent, 3)
                        insert_technics('pogr', 58, row)
                    if int(row[0]) == 30:
                        self.changes(month[0], 'Погрузчики', 'Погрузчик Shantui SL60W №30', percent, 3)
                        insert_technics('pogr', 59, row)
                    if int(row[0]) == 25:
                        self.changes(month[0], 'Погрузчики', 'Погрузчик Volvo L120G №25', percent, 3)
                        insert_technics('pogr', 60, row)

            wb.save(f'{path}\PO6_year_{p_date}-{p_time}.xlsx')
            self.timer_stop()
            os.system(f'start excel.exe "{path}\PO6_year_{p_date}-{p_time}.xlsx"')
        except Exception as e:
            messagebox.showerror('Ошибка', 'Выберите хотя бы 1 месяц')
            print(e)
            return
